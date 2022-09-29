import logging
from openpyxl import load_workbook
logger = logging.getLogger(__name__)


def filter_empty(old_l):
    """过滤序列中的空值"""
    new_l = []
    for i in old_l:
        if i:
            new_l.append(i)
    return new_l


def data_by_excel(file):
    wb = load_workbook(file)
    logger.info(f'文件{file=},包含了{len(wb.worksheets)}sheet')
    suite_dict = {}  # 以套件名称为key，以用例为value

    for ws in wb.worksheets:
        case_dict = {}  # 以名称为key，以步骤为value的字典
        case_name = ""

        for line in ws.iter_rows(values_only=True):
            _id = line[0]
            logger.debug(f"正在处理下一行 :{line}")

            if isinstance(_id, int):
                if _id == -1:  # 用例名称
                    case_name = line[3]
                    case_dict[case_name] = []  # 以用例名称为key，创建新的空用例
                elif _id > 0:  # 步骤
                    case_dict[case_name].append(filter_empty(line))  # 为用例填充步骤`

        logger.info(f"Sheet {ws.title},包含了{len(case_dict)}个用例")
        suite_dict[ws.title] = case_dict  # 本测试套件的所有用例

    logger.debug(f"加载测试用例用例 {suite_dict=}")
    return suite_dict